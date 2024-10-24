import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 读取 Excel 文件
file1 = 'updated_table2.xlsx'
file2 = 'table3.xlsx'

df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

# 假设 SKU 编码的列名为 'SKU编码'
# 将 df1 中的相关列与 df2 合并
merged_df = df2.merge(df1[['SKU编码', '小单位条形码_new', '商品名称_new', '单价_new']],
                      on='SKU编码', how='left',
                      suffixes=('', '_new'))

# 保存合并后的 DataFrame 到一个新的 Excel（临时保存）
merged_df.to_excel("merged_file.xlsx", index=False)

# 使用 openpyxl 加载刚刚保存的文件以进行高亮操作
wb = load_workbook("merged_file.xlsx")
ws = wb.active

# 定义填充格式
highlight = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# 高亮显示 SKU编码 和 SKU编码_new 相等的行
for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
    sku = row[df2.columns.get_loc('SKU编码')].value
    sku_new = row[df2.columns.get_loc('SKU编码') + 4].value # 根据插入位置调整索引
    if sku == sku_new:
        for cell in row:
            cell.fill = highlight

# 保存最终结果
final_file = "highlighted_table3.xlsx"
wb.save(final_file)

print(f"处理完成，结果保存在 {final_file}")
