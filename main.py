import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

file1 = 'table1.xls'
file2 = 'table2.xlsx'

# 使用Pandas读取Excel表格
df1 = pd.read_excel(file1, sheet_name=0, engine='xlrd')
df2 = pd.read_excel(file2, sheet_name=0, engine='openpyxl')

# 验证并打印列名
print("Table 1 Columns:", df1.columns)
print("Table 2 Columns:", df2.columns)

# 根据实际列名替换
cols_barcode_1 = '小单位条形码'  # 表1的小单位条形码的实际列名
cols_barcode_2 = '商品条码'     # 表2的商品条码的实际列名

# 确认df1中存在的实际列名，假设为 'Barcode', 'ProductName', 'Price'
barcode_col_1 = '小单位条形码'
product_name_col_1 = '商品名称'
price_col_1 = '单价'

# 将两个数据框中的条形码列都转换为字符串
df1[cols_barcode_1] = df1[cols_barcode_1].astype(str)
df2[cols_barcode_2] = df2[cols_barcode_2].astype(str).str.lstrip("'")

# 在特定位置插入新列
insert_col_position = df2.columns.get_loc(cols_barcode_2) + 1

new_data = df2.copy()
new_data.insert(insert_col_position, '小单位条形码_new', '')
new_data.insert(insert_col_position + 1, '商品名称_new', '')
new_data.insert(insert_col_position + 2, '单价_new', '')

for index, row in df2.iterrows():
    matching_rows = df1[df1[cols_barcode_1] == row[cols_barcode_2]]
    if not matching_rows.empty:
        matched_row = matching_rows.iloc[0]
        new_data.loc[index, '小单位条形码_new'] = matched_row[barcode_col_1]
        new_data.loc[index, '商品名称_new'] = matched_row[product_name_col_1]
        new_data.loc[index, '单价_new'] = matched_row[price_col_1]

output_file = 'updated_table2.xlsx'
new_data.to_excel(output_file, index=False)

wb = load_workbook(output_file)
ws = wb.active

green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type="solid")

# 检查和填充颜色时也应该用字符串比较
for row in ws.iter_rows(min_row=2, min_col=insert_col_position+1, max_col=insert_col_position+3):
    for cell in row:
        barcode = cell.value
        # 确保条形码是作为字符串处理的
        if barcode and str(barcode) in df1[cols_barcode_1].values:
            corresponding_row = df1[df1[cols_barcode_1] == str(barcode)].iloc[0]
            if str(barcode) == corresponding_row[barcode_col_1]:
                cell.fill = green_fill

wb.save(output_file)

print(f"Updated file saved as {output_file}.")
